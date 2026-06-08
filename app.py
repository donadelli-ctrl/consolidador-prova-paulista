import streamlit as st
import pandas as pd
import zipfile
import tempfile
import os
import re
import unicodedata
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(
    page_title="Consolidador Prova Paulista",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Consolidador Prova Paulista - Oficial 1.0")

nome_escola = st.text_input(
    "🏫 Nome da Escola"
)


def normalizar(txt):

    if txt is None:
        return ""

    txt = str(txt)

    txt = unicodedata.normalize(
        "NFKD",
        txt
    ).encode(
        "ASCII",
        "ignore"
    ).decode(
        "utf-8"
    )

    txt = txt.upper()

    txt = re.sub(
        r"\s+",
        " ",
        txt
    )

    return txt.strip()


def turma_curta(texto):

    texto = normalizar(texto)

    numero = re.search(r"(\d+)", texto)

    letra = re.search(
        r"\b([A-Z])\b",
        texto
    )

    if numero and letra:
        return f"{numero.group(1)}{letra.group(1)}"

    return texto


def status(valor):

    if valor < 0.50:
        return "Abaixo do Básico"

    elif valor < 0.70:
        return "Básico"

    elif valor < 0.90:
        return "Adequado"

    return "Proficiente"


def cor_status(txt):

    cores = {

        "Abaixo do Básico": "F4CCCC",
        "Básico": "FFF2CC",
        "Adequado": "D9EAD3",
        "Proficiente": "CFE2F3"

    }

    return cores.get(
        txt,
        "FFFFFF"
    )


xlsm = st.file_uploader(
    "📎 Planilha XLSM",
    type=["xlsm"]
)

zip_file = st.file_uploader(
    "📦 Arquivo ZIP",
    type=["zip"]
)


if st.button("🚀 Gerar Consolidado"):

    if not nome_escola.strip():

        st.error(
            "Informe o nome da escola."
        )

        st.stop()

    if not xlsm:

        st.error(
            "Selecione o XLSM."
        )

        st.stop()

    if not zip_file:

        st.error(
            "Selecione o ZIP."
        )

        st.stop()

    resultados = {}

    with tempfile.TemporaryDirectory() as pasta:

        caminho_zip = os.path.join(
            pasta,
            zip_file.name
        )

        with open(
            caminho_zip,
            "wb"
        ) as f:

            f.write(
                zip_file.getbuffer()
            )

        with zipfile.ZipFile(
            caminho_zip,
            "r"
        ) as z:

            z.extractall(pasta)

        arquivos_excel = []

        for raiz, _, arquivos in os.walk(pasta):

            for arq in arquivos:

                if arq.lower().endswith(
                    ".xlsx"
                ):

                    arquivos_excel.append(
                        os.path.join(
                            raiz,
                            arq
                        )
                    )

        for arquivo in arquivos_excel:

            try:

                turma = os.path.splitext(
                    os.path.basename(
                        arquivo
                    )
                )[0]

                turma = normalizar(
                    turma
                )

                df = pd.read_excel(
                    arquivo
                )

                if "Nome" not in df.columns:
                    continue

                for _, row in df.iterrows():

                    nome = normalizar(
                        row["Nome"]
                    )

                    resultados[
                        (
                            turma,
                            nome
                        )
                    ] = {

                        "LP": row["PORT"],
                        "MAT": row["MAT"]

                    }

            except:
                pass

    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".xlsm"
    ) as tmp:

        tmp.write(
            xlsm.getbuffer()
        )

        caminho_xlsm = tmp.name

    wb = load_workbook(
        caminho_xlsm,
        keep_vba=True
    )

    encontrados = 0
    nao_encontrados = 0
    lp_preenchidos = 0
    mat_preenchidos = 0
    total_alunos = 0

    for ws in wb.worksheets:

        for linha in range(
            4,
            ws.max_row + 1
        ):

            estudante = ws.cell(
                linha,
                2
            ).value

            if not estudante:
                continue

            total_alunos += 1

            turma = turma_curta(
                ws.cell(
                    linha,
                    1
                ).value
            )

            nome = normalizar(
                estudante
            )

            chave = (
                turma,
                nome
            )

            if chave not in resultados:

                nao_encontrados += 1
                continue

            reg = resultados[
                chave
            ]

            lp = reg["LP"]
            mat = reg["MAT"]

            if pd.notna(lp):

                ws.cell(
                    linha,
                    7
                ).value = lp

                ws.cell(
                    linha,
                    7
                ).number_format = "0.0%"

                st_lp = status(
                    float(lp)
                )

                ws.cell(
                    linha,
                    8
                ).value = st_lp

                ws.cell(
                    linha,
                    8
                ).fill = PatternFill(
                    "solid",
                    fgColor=cor_status(
                        st_lp
                    )
                )

                lp_preenchidos += 1

            else:

                ws.cell(
                    linha,
                    8
                ).value = ""

            if pd.notna(mat):

                ws.cell(
                    linha,
                    9
                ).value = mat

                ws.cell(
                    linha,
                    9
                ).number_format = "0.0%"

                st_mat = status(
                    float(mat)
                )

                ws.cell(
                    linha,
                    10
                ).value = st_mat

                ws.cell(
                    linha,
                    10
                ).fill = PatternFill(
                    "solid",
                    fgColor=cor_status(
                        st_mat
                    )
                )

                mat_preenchidos += 1

            else:

                ws.cell(
                    linha,
                    10
                ).value = ""

            encontrados += 1

    saida_xlsm = BytesIO()

    wb.save(
        saida_xlsm
    )

    saida_xlsm.seek(0)

st.success(
    f"🏫 Escola: {nome_escola}"
)

st.subheader(
    "📊 Relatório"
)

c1, c2, c3, c4, c5 = st.columns(5)

   c1.metric(
    "👨‍🎓 Total",
    total_alunos
)

    c2.metric(
        "Encontrados",
        encontrados
    )

    c3.metric(
        "Não Encontrados",
        nao_encontrados
    )

    c4.metric(
        "LP",
        lp_preenchidos
    )

    c5.metric(
        "MAT",
        mat_preenchidos
    )

    nome_arquivo = normalizar(
        nome_escola
    ).replace(
        " ",
        "_"
    )

    st.download_button(
        "📥 Baixar XLSM",
        data=saida_xlsm.getvalue(),
        file_name=f"{nome_arquivo}_CONSOLIDADO.xlsm",
        mime="application/vnd.ms-excel"
    )

st.divider()

if st.button("🔄 Limpar Tudo"):

    st.session_state.clear()

    st.rerun()
